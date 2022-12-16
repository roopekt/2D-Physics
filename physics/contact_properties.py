import openpyxl
from dataclasses import dataclass
from itertools import count
import os.path as path

@dataclass
class ContactProperties:
    elasticity: float = 0.3
    rest_friction_coefficient: float = 0.3
    dynamic_friction_coefficient: float = 0.2

    @classmethod
    def from_string(cls, string):
        numbers = [float(e) for e in string.split(",")]
        assert len(numbers) == 3, f"Wrong number of numbers in cell. Expected 3, Actual {len(numbers)}"

        return cls(
            elasticity = numbers[0],
            rest_friction_coefficient = numbers[1],
            dynamic_friction_coefficient = numbers[2]
        )

class ContactTable:

    def __init__(self):
        self.materials = {}

        file_path = path.join(path.dirname(__file__), "contact_property_table.xlsx")
        with openpyxl.load_workbook(filename=file_path) as file:
            table = file.tables.values()[0]

            column_names = []
            for i in count(start=2):
                new_name = table.cell(row=1, column=i).value
                if new_name == "":
                    break
                else:
                    column_names.append(new_name)

            row_names = []
            for i in count(start=2):
                new_name = table.cell(row=i, column=1).value
                if new_name == "":
                    break
                else:
                    row_names.append(new_name)

            for i, row_name in enumerate(row_names):
                for i, column_name in enumerate(column_names):
                    cell_value = table.cell(row=i+2, column=i+2).value
                    self.materials[(row_name, column_name)] = ContactProperties.from_string(cell_value)
                    self.materials[(column_name, row_name)] = self.materials[(row_name, column_name)]

    def get_contact_properties(self, material_name_A, material_name_B):
        return self.materials[(material_name_A, material_name_B)]
